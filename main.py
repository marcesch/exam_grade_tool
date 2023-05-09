# This is a sample Python script.
import csv
import logging
import os

import openpyxl

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


from exam import Exam, Student
from classes import Class
# from student import Student
from category import Category

"""
TODO keep list of classes to index everything -- in GUI, user can then select which class to edit
TODO add functionality to delete old classes
TODO upon startup, load everything from csv
=> sample run:
1. Load list of classes from disk (stored as csv)
2. Let user choose class to be viewed / edited
3. Load class-list, list of exams and exams themselves from stored csv
=> either just load all exams found in a folder or keep index (stored somewhere as csv), probably the latter as I also 
need to store information about the category weights / ...

=> Probably need to create new file for loading / storing stuff. Keep it minimal, loading takes quite a long time...
TODO don't always store everything on disk, only on major changes and on shutdown!

TODO add fucntionality to create exam summary as a PDF or so (average, mean, quartils, standard deviation, ..)
"""

FOLDERPATH = "./tmp/klassen/"


class Overview:
    """
    keep list of classes, ...

    """

    def __init__(self):
        self.classes = []

    def load_categories_and_exams(self, class_obj: Class, path_folder_exams: str = None):
        """
        :return:
        """

        # TODO caller can take care of panic mode -- if this function returns an error, the caller invokes this function again with the backup-filepath

        if path_folder_exams == None:
            dirpath = class_obj.filename_base_exam
        else:
            dirpath = path_folder_exams

        # TODO include checks on validity / existence of folder

        # Open workbook
        wb = openpyxl.load_workbook(dirpath)
        sheet = wb.active

        # Find row containing "Nachname"
        # TODO I think this is wrong with max_row, using 20 instead.
        # for row in sheet.iter_rows(min_row=1, max_row=1):
        row_index = -1
        for row in sheet.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value == "Nachname":
                    row_index = cell.row

        if row_index == -1:
            raise RuntimeWarning(
                f"Could not locate the string Nachname in the first row. Please make sure that you did not alter the contents of that file")

        # TODO insert checks if things go wrong
        # Iterate over category rows
        for row in sheet.iter_rows(min_row=2, max_row=row_index - 5, min_col=2, max_col=4):
            category_name = row[0].value
            weight = float(row[1].value)
            grading_type = row[2].value
            category = Category(category_name, class_obj.term, class_obj.name, weight, grading_type)
            class_obj.categories.append(category)

        # Now read in the contents for the categories => iterate over all columns that contain data for exams
        col_index = 3

        while True:
            # read contents for current exam
            cell_min_grade = sheet.cell(row=row_index - 4, column=col_index)
            cell_max_grade = sheet.cell(row=row_index - 4, column=col_index + 1)
            cell_max_points = sheet.cell(row=row_index - 3, column=col_index)
            cell_pts_for_max = sheet.cell(row=row_index - 3, column=col_index + 1)
            cell_cmp_mode = sheet.cell(row=row_index - 2, column=col_index)
            cell_exam_name = sheet.cell(row=row_index - 1, column=col_index)
            cell_exam_cat = sheet.cell(row=row_index - 1, column=col_index + 1)

            # check if there still is data for exam in this column
            if cell_exam_name.value is None:
                break

            # create new exam and add it to correct category
            min_grad = float(cell_min_grade.value)
            max_grade = float(cell_max_grade.value)
            max_points = int(cell_max_points.value)
            pts_for_max = int(cell_pts_for_max.value)
            computation_mode = cell_cmp_mode.value
            exam_name = cell_exam_name.value
            exam_cat = cell_exam_cat.value

            # read in  student data:
            grades_exam = {}
            points_exam = {}
            for row in sheet.iter_rows(min_row=row_index + 1):
                firstname = row[0].value
                lastname = row[0].value
                student = class_obj.get_student(firstname, lastname)
                if row[col_index] != None:
                    points_exam[student] = int(row[col_index])
                if row[col_index + 1] != None:
                    grades_exam[student] = float(row[col_index + 1])

            # create exam object
            class_obj.add_exam(exam_name,
                               exam_cat,
                               max_points,
                               pts_for_max,
                               min_grade=min_grad,
                               max_grade=max_grade,
                               achieved_points=points_exam,
                               achieved_grades=grades_exam,
                               computation_mode=computation_mode)

            col_index += 2

    # def load_class_data(self, path_classlist=None):
    #     """
    #     Loads the following fields from the file "klassen" -> [year]_[term]_[name].csv:
    #     - year = [year]
    #     - term = [term]
    #     - name = [name]
    #     - students from contents of file
    #     => maybe don't load year, term, name from student -- they will be initialized when class is created anyway..let
    #     the caller take care of that
    #     :param filepath: can overwrite filepath / which folder should be taken into consideration for reading in the classlists
    #     :return:
    #     """
    #     # TODO caller can take care of panic mode -- if this function returns an error, the caller invokes this function again with the backup-filepath
    #
    #     # TODO ensure that old classlists get moved to /archive or so, otherwise they always get loaded as active class
    #
    #     # TODO also load report id
    #
    #     if path_classlist == None:
    #         dirpath = FOLDERPATH
    #     else:
    #         # TODO maybe include some checks on validity of path
    #         dirpath = path_classlist
    #
    #     for file in os.listdir(dirpath):
    #         if file.endswith(".csv"):
    #             year, term, name = file[:-4].split("_")
    #             class_obj = Class(name, term, int(year))
    #             with open(os.path.join(dirpath, file), 'r') as f:
    #                 lines = f.readlines()
    #                 header = lines[0].strip().split(",")
    #                 if header[0] != "Nachname" or header[1] != "Vorname":
    #                     raise RuntimeWarning(
    #                         "First row of CSV file should contain the strings 'Nachname' and 'Vorname'")
    #                 for line in lines[1:]:
    #                     lastname, firstname = line.strip().split(",")
    #                     student = Student(firstname, lastname)
    #                     class_obj.students.append(student)
    #             self.classes.append(class_obj)

    def load_classes(self):
        """
        loads list of classes from memory
        => how to deal with multiple intstances of classes? => mb load all of them, let user decide which to pick.

        => when does it get called? on startup, probably?
        :return:
        """

        # Loop through all files in the directory
        for filename in os.listdir(FOLDERPATH):
            # Check if the file has a .csv extension
            if filename.endswith(".csv"):
                print(f"Extracting data from filename {filename}")
                year, term, name = filename[:-4].split("_", 2)
                class_obj = Class(name, term, int(year))
                print(f"Trying to open {filename}")
                try:
                    with open(os.path.join(FOLDERPATH, filename), 'r') as f:
                        lines = f.readlines()
                        header = lines[0].strip().split(",")
                        if header[0] != "Nachname" or header[1] != "Vorname":
                            raise RuntimeWarning(
                                "First row of CSV file should contain the strings 'Nachname' and 'Vorname'")
                        for line in lines[1:]:
                            lastname, firstname = line.strip().split(",")
                            print(f"Got here with {firstname}, {lastname}")
                            student = Student(firstname, lastname)
                            class_obj.students.append(student)
                    self.classes.append(class_obj)

                    # get report_id (to make sure not to create reports multiple times)
                    max_report_no = 0
                    for file in os.path.join(FOLDERPATH, filename[:-4]):
                        # TODO test this functionality
                        if "report" in file:
                            try:
                                if int(file[file.find("report"):file.find("report") + 2]) > max_report_no:
                                    max_report_no = int(file[file.find("report"):file.find("report") + 2])
                            except:
                                logging.warning(f"Could not extract report number from filename {file}")
                    class_obj.report_id = max_report_no

                except:
                    if class_obj != None:
                        logging.error(f"Could not create class from {filename} for class {class_obj}. \n"
                                      f"Make sure that the file is named correctly (YEAR_TERM_NAME) and does contain a list of students.\n"
                                      f"Make sure, the first row of the Row contains a field Nachname")
                    else:
                        logging.error(f"Could not create class from {filename}. \n"
                                      f"Make sure that the file is named correctly (YEAR_TERM_NAME) and does contain a list of students.\n"
                                      f"Make sure, the first row of the Row contains a field Nachname")
