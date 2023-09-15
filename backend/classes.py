import csv
import typing
import os
import logging
import shutil
import openpyxl
import time

from backend.category import *
from backend.exam import *
from backend.student import Student



""""
TODO: sanitize / ... students names... what if there are sume ugly shit thigns like accents etc?
=> might have to slightly alter the contains_student function etc. (always compare sanitized version)

=> add possibility to import old examples (e.g. for categories etc., otherwise have to do that everytime again)


TODO: when should I store stuff?

TODO cna python also handle xlsx files? might be a little bit more convenient than using csv... e.g. when creating report

"""


class Class:
    def __init__(self, name: str, term: str, year: int):
        self.name: str = name
        self.year: int = year

        self.students: list[Student] = []
        self.categories: list[BaseCategory] = []
        self.exams: list[Exam] = []
        # keeps a list of current average_grades for each student (so it can be displayed in GUI)
        self.average_grades: dict[Student, float] = {}

        if (term.lower() == "hs"):
            self.term: str = term.upper()
        elif (term.lower() == "fs"):
            self.term: str = term.upper()
        else:
            raise RuntimeError(f"[CLASS] Either choose HS or FS for term (got {term})")

    ####################### FIELD MANIPULATIONS INTERFACE #######################

    def update_name(self, new_name: str):
        self.name = new_name

    def change_category_of_exam(self, exam: Exam, new_category: BaseCategory):
        """
        Sets another category for the exam
        :param exam: exam whose cateogry should be changed
        :param new_category: new category for the exam
        :return:
        """

        exam.change_category(new_category)
        if not self.contains_category(new_category):
            self.categories.append(new_category)

    def add_multiple_students(self, students: list[tuple[str, str]]):
        """
        Adds students to class, based on received list of (Firstname, Lastname)
        :param students:  list of (firstname, lastname)-tuples
        :return: list of students that could not be added
        """

        # TODO callers need to check for list being empty

        list_of_failed_students = []
        for student_tuple in students:
            firstname = student_tuple[0]
            lastname = student_tuple[1]
            try:
                self.add_student(firstname, lastname)
            except RuntimeError as e:
                list_of_failed_students.append(student_tuple)

        if len(list_of_failed_students) != 0:
            logging.warning(f"[Class] Could not add the following students: {list_of_failed_students}")

        return list_of_failed_students

    def add_student(self, firstname, lastname):
        """
        Add single student (based on firstname / lastname) to class
        :return:
        :raises: RuntimeError if student is already in class
        """
        if self.contains_student(firstname, lastname):
            raise RuntimeError(
                f"[CLASS] Cannot add student {firstname} {lastname} a second time. Consider adding suffix (John Doe2)")
        self.students.append(Student(firstname, lastname))


    def delete_student(self, student: Student):
        """
        Deletion based on student object instead of name. Probably gonna use that function consitently #TODO
        :param student:
        :return:
        """
        if student in self.students:
            self.students.remove(student)
        else:
            raise RuntimeError(f"Could not find student {student} in class {self}")

    def delete_exam(self, exam: Exam):
        """
        Deletion based on object reference instead of name. Replace other funciton usages with delete instead of remove
        :param exam:
        :return:
        """
        if exam in self.exams:
            self.exams.remove(exam)
        else:
            raise RuntimeError(f"Could not find exam {exam} in class {self}")

    def delete_category(self, category: BaseCategory):
        """
        Removes the category from this class.
        :param category:
        :raises: RuntimeError if category is not known or there still are exams of that category.
        """

        if self.contains_category(category):
            if self.number_exams_of_category(category) != 0:
                raise RuntimeError(f"[CLASS] Cannot delete category. There are still {self.number_exams_of_category(category)} exams in my list of that category!")
            self.categories.remove(category)
        else:
            raise RuntimeError(f"Could not find that category in the list {self.categories}")


    def add_category(self,
                     category_name: str,
                     category_type: str,
                     weight: float = 0,
                     bonus_amount: float = 0,
                     number_drop_grades: int = 0):
        """
        Add category. Based on the key "type", the exact type of the category should be distinguished (e.g. CategoryDefault, ..)
        :param category_description: dictionary containing the necessary fields to create category.
        :return:
        """

        type = category_type
        new_category = None
        if type == "default":
            new_category = CategoryDefault(category_name, weight)
        elif type == "drop_grade":
            new_category = CategoryWithDroppedGrades(category_name, weight, number_drop_grades)
        elif type == "bonus":
            new_category = CategoryBonus(category_name, bonus_amount)
        else:
            raise RuntimeError(f"[CLASS] Cannot create category of type {type}")

        self.categories.append(new_category)


    ################# SMALL UTILITY ##################

    def __str__(self):
        return f"{self.name} {self.term.upper()}{self.year}"

    def __repr__(self):
        return f"C-{self.name} {self.term.upper()}{self.year}"

    # FUNCTIONS TO ENABLE EASY SORTING

    def __le__(self, other):
        """
        Comparison based on year and term
        HS2020 -> FS2020 -> HS2021 -> ..
        Tie-breaker: alphanumerical order of name
        """
        if self.term == other.term and self.year == other.year:
            return self.name <= other.name
        elif self.year == other.year:
            return self.term.lower() == "hs"
        else:
            return self.year <= other.year

    def __lt__(self, other):
        if self.term == other.term and self.year == other.year:
            return self.name < other.name
        elif self.year == other.year:
            return self.term.lower() == "hs"
        else:
            return self.year < other.year

    def __eq__(self, other):
        "Does NOT compare exams etc. ==> only used for sorting!!"
        return self.term == other.term and self.year == other.year and self.name == other.name


    def number_exams(self):
        """
        :return: number of currently stored exams
        """
        return len(self.exams)

    def number_exams_of_category(self, category: BaseCategory):
        """
        Returns the number of exams that match the category
        :param cat:
        :return:
        """

        return len(self.get_exams_of_category(category))


    def number_exams_for_student(self, student: Student):
        """
        Returns number of exams that the student has written
        :param student:
        :return:
        """

        return len(self.get_exams_for_student())

    def number_categories(self):
        return len(self.categories)


    def contains_student(self, firstname, lastname):
        """

        :param firstname:
        :param lastname:
        :return: True iff a student with firstname / lastname is in self.students
        """

        try:
            self.get_student(firstname, lastname)
        except RuntimeError as e:
            return False
        return True

    def contains_category(self, category_name: str):
        try:
            self.get_category(category_name)
        except RuntimeError as e:
            return False
        return True

    def contains_exam(self, exam_name: str):
        try:
            self.get_exam(exam_name)
        except RuntimeError as e:
            return False
        return True

    def get_student(self, firstname, lastname):
        """
        Find student object, based on either first or lastname (or both)
        :param first:
        :param last:
        :return: Reference to Student object corresponding to the firstname / lastname combination
        """

        if firstname == "" and lastname == "":
            raise RuntimeError(f"[CLASS] Must give valid name to find student")
        firstname = firstname.lower()
        lastname = lastname.lower()
        for student in self.students:
            if student.firstname == firstname and student.lastname == lastname:
                return student

        raise RuntimeError(f"[CLASS] ould not find student {firstname} {lastname} in my list")

    def get_category(self, category_name: str):
        """
        Get category object corresponding to category_name
        :param category_name:
        :return:
        """

        for cat in self.categories:
            if cat.name == category_name:
                return cat

        raise RuntimeError(f"[CLASS] No category {category_name} found in my list: \n {self.categories}")


    def get_exam(self, exam_name: str):
        """
        Gets exam object based on name
        :return: reference to Exam object corresponding to the name
        """
        for ex in self.exams:
            if ex.name == exam_name:
                return ex

        raise RuntimeError(f"Could not find exam {exam_name} in list {self.exams}")

    def get_exams_of_category(self, category: BaseCategory):
        """
        Returns list of all exams of a given category
        :param category:
        :return:
        """

        exams = []
        for exam in self.exams:
            if exam.category == category:
                exams.append(exam)

        return exams

    def get_exams_for_student(self, student: Student):
        """
        Returns list of exams that the student was part of
        :param student:
        :return:
        """

        return [exam for exam in self.exams if student in exam.grades]


    ################## IMPORTANT FUNCTIONALITY #####################


    def add_exam(self,
                 exam_name: str,
                 category: BaseCategory,
                 type: str="default",
                 additional_args={}
                 ):
        """
        Provides functionality to add exams to this class
        :param exam_name: Name of exam
        :param category: Under which category this exam should be put
        :param type: Type of the exam (see Exam class)
        :param additional_args: used to initialize different kinds of exams
        :return: reference to newly created object
        """

        if self.contains_exam(exam_name):
            raise RuntimeError(f"[CLASS] Can not create exam {exam_name}: There is an exam with that name")

        if not category in self.categories:
            self.categories.append(category)

        if type == "linear" or type == "default":
            # create Exam mode linear
            try:
                # unpack things from additional args
                max_points = additional_args["max_points"]
                # create exam object, potentially with non-required args

                # TODO continue here
                # => how to deal with non-required arguments?
                # a) simply set defaults a second time, but that's pretty messy
                # b) hardcode all combinations of cases

                point_for_max = 0
                if "points_for_max" not in additional_args:
                    logging.warning(f"[CLASS] Did not receive point_for_max argument. Defaulting to max points")
                    point_for_max = max_points
                else:
                    point_for_max =additional_args["points_for_max"]


                exam_obj = ExamModeLinear(exam_name,
                                          self.term,
                                          self.year,
                                          self.name,
                                          category,
                                          max_points,
                                          point_for_max,
                                          additional_args)

                self.exams.append(exam_obj)
                return exam_obj

            except KeyError as e:
                raise RuntimeError(f"[CLASS] Could not unpack arguments for this exam type: \n {e.__str__()}")

        elif type == "linear_with_pass":
            try:
                max_points = additional_args["max_points"]
                points_for_pass = additional_args["points_for_pass"]
                points_for_max = 0

                if "points_for_max" not in additional_args:
                    logging.warning(f"[CLASS] Did not receive point_for_max argument. Defaulting to max points")
                    point_for_max = max_points
                else:
                    point_for_max = additional_args["points_for_max"]

                exam_obj = ExamModeLinearWithPassingPoints(exam_name,
                                                           self.term,
                                                           self.year,
                                                           self.name,
                                                           category,
                                                           max_points,
                                                           points_for_max,
                                                           points_for_pass,
                                                           additional_args)
                self.exams.append(exam_obj)
                return exam_obj
            except KeyError as e:
                raise RuntimeError(f"[CLASS] Could not unpack arguments for this exam type: \n {e.__str__()}")

        elif type == "fixed_point_scheme":
            raise  NotImplementedError
        elif type == "manually" or "direct_grade":

            # TODO take care of additional arguments more thoroughly, but I need this to run for testing...

            exam_obj = ExamModeSetGradeManually(exam_name, self.term, self.year, self.name, category, additional_args)
            self.exams.append(exam_obj)
            return exam_obj

            raise NotImplementedError
        else:
            raise RuntimeError(f"[CLASS] Don't know that exam type")



    def edit_category(self, category: BaseCategory, properties):
        """
        Allows altering the category type, properties, ...
        :param category:
        :param properties: dict containing necessary information to alter category.
        :return:
        """
        # TODO make sure to change all exams of that category accordingly
        raise NotImplementedError


    def compute_average_grade_student(self, student: Student, DEBUG=False):
        """
        Computes average grade of student across all exams
        :param student:
        :return:
        """

        """
        TODO Issues with this function:
        - Very unelegant solution
        - Take voluntary flag into consideration for the computation of the grade
        - Find better solution for weight != 1
        """

        # TODO there are bugs here -- simply rewrite everything, it's super ugly anyway

        # Collect all exams that the student has written, by category

        exams_by_category = {}

        bonus_exams = []
        for exam in self.exams:
            if isinstance(exam.category, CategoryBonus):
                # ignore the bonus grades, they are added at the end
                bonus_exams.append(exam)
                continue
            if student in exam.grades:
                if exam.category in exams_by_category:
                    exams_by_category[exam.category].append(exam)
                else:
                    exams_by_category[exam.category] = [exam]


        # TODO how to deal with this case -- ask Selina
        # check that student has written an exam in at least every category.
        if len(exams_by_category) != len(self.categories):
            raise RuntimeError(f"[CLASS] Error: student did not write exams from all categories: {exams_by_category.keys} (should be {self.categories})")

        # TODO mb use more elegant algorithm -- I think I'm actually doing a submodular set function here
        def generate_subsets(elements):
            n = len(elements)
            for i in range(1 << n):  # Iterate from 0 to 2^n - 1
                subset = [elements[j] for j in range(n) if (i & (1 << j)) != 0]
                yield subset

        resulting_grades = {}
        for cat in exams_by_category:
            mandatory_exam_grades = [exam.grades[student] for exam in exams_by_category[cat] if not exam.voluntary]
            voluntary_exams = [exam.grades[student] for exam in exams_by_category[cat] if exam.voluntary]
            possible_permutations = generate_subsets(voluntary_exams)
            max_value = 1
            for subset in possible_permutations:
                current_set = mandatory_exam_grades + subset
                averaged_grade = cat.aggregate_grades(current_set)
                max_value = max(max_value, averaged_grade)
            resulting_grades[cat] = max_value

        if DEBUG:
            print(f"[CLASS, DEBUG] Aggregated averaged grades for {student} with voluntary exams:\n{resulting_grades}")

        # compute the weighted average over all weighted categories
        sum_weigth = sum([cat.weight for cat in resulting_grades])
        if sum_weigth == 0:
            raise RuntimeError(f"Can't compute grade without any weights for the categories")
        elif sum_weigth != 1:
            logging.warning(f"[CLASS] Weights don's sum up to 1. I rescale the categories accordingly...")
            if DEBUG:
                print(f"[CLASS, DEBUG] Weights don's sum up to 1. I rescale the categories accordingly...")
                print(f"[CLASS, DEBUG] Categories, weights: {sum_weigth}, {[(cat.name, cat.weight) in resulting_grades]}")

        weighted_grade = 0
        for cat in resulting_grades:
            weighted_grade += 1/sum_weigth * cat.weight * resulting_grades[cat]

        if DEBUG:
            print(f"[CLASS, DEBUG] Got grade {weighted_grade} for {student} (without bonus)")

        # now, compute the resulting grade by also considering the bonus grades
        final_grade = weighted_grade
        for exam in bonus_exams:
            raise RuntimeError(f"[CLASS] Don't support bonus grades as of now.")

        # ensure grade boundaries
        min_grade = 1
        max_grade = 6

        return max(min_grade, min(max_grade, final_grade))

    def compute_average_grades(self):
        """
        Computes average grades of all students
        :return:
        """

        for student in self.students:
            try:
                self.average_grades[student] = self.compute_average_grade_student(student)
            except:
                logging.warning(f"[CLASS] Could not compute average grade for student {student}")
                # mark missing average grade for those students
                self.average_grades[student] = -1

    def export_grade_report(self, filepath = ""):
        """
        Merge the old functions here
        :param filepath:
        :return:
        """
        raise NotImplementedError

    def export_class_overview(self, filetype="xlsx", filepath=""):
        """
        Stores all exams to disk. Most important function, as this allows manual overwrite
        See the layout_gradefiles for details on what goes where, ...
        :param filetype:
        :param filepath: use default path if none is given
        :return:
        """

        # TODO merge with create_grade_report

        raise NotImplementedError

        # TODO sort students according to last name
        # sort students according to lastname
        # TODO can also support firstname, theoretically
        self.students = sorted(self.students, key=lambda student: student.lastname)

        # sanitize file type
        # if filetype not in self.supported_filetypes:
        if filetype != "xlsx":
            logging.warning(f"Don't know that filetype {filetype}. Using xlsx instead")
            filetype = "xlsx"

        output_name = os.path.join(self.filename_base_exam, "pruefungen.xlsx")
        # check if an old exam with that name already exists - if yes, move it to trash
        # logging.info(f"Storing {exam.name}_{exam.term} to {output_name}.{filetype}")
        # print(f"Storing {exam.name}_{exam.term} to {output_name}.{filetype}")

        # if os.path.exists(output_name):
        #     logging.warning("Found old file containing the exams. Moving it to the .trash folder")
        #     shutil.move(output_name, f"{self.folder_shadow}{time.time()/60}_{output_name}")

        # store the exam to disk, given the chosen mode
        # TODO also store max points, points for 6 in exam
        if filetype == "xlsx":
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # reserve some space on top of Excel file (for class name, weights of categories, ...)
            # For loading, search for "Nachname" or so to get this index.

            # set number of headers, accoring to layout in .ods file
            amount_header_columns = 1 + 1 + len(self.categories) + 5

            worksheet.cell(row=1, column=1, value="Klassenname")
            worksheet.cell(row=1, column=2, value=self.name)

            worksheet.cell(row=2, column=1, value="Categories")
            worksheet.cell(row=2, column=2, value="Category name")
            worksheet.cell(row=2, column=3, value="Category weight")
            worksheet.cell(row=2, column=4, value="Grading type")
            for i, cat in enumerate(self.categories):
                worksheet.cell(row=2 + i, column=2, value=cat.name)
                worksheet.cell(row=2 + i, column=3, value=cat.weight)
                worksheet.cell(row=2 + i, column=4, value=cat.grading_type)
                worksheet.row_dimensions[2 + i].hidden = True

            worksheet.cell(row=amount_header_columns - 1, column=1, value="Nachname")
            worksheet.cell(row=amount_header_columns - 1, column=2, value="Vorname")
            for row, student in enumerate(self.students, start=amount_header_columns):
                worksheet.cell(row=row, column=1, value=student.lastname)
                worksheet.cell(row=row, column=2, value=student.firstname)

            column_iterator = 0
            for cat_column, cat in enumerate(self.categories):
                for exam_column, exam in enumerate(cat.exams):
                    # store exam headers
                    exam_column = 2 * column_iterator + 3
                    column_iterator = column_iterator + 1
                    worksheet.cell(row=amount_header_columns - 5, column=exam_column, value=exam.min_grade)
                    worksheet.cell(row=amount_header_columns - 5, column=exam_column + 1, value=exam.max_grade)
                    worksheet.cell(row=amount_header_columns - 4, column=exam_column, value=exam.max_points)
                    worksheet.cell(row=amount_header_columns - 4, column=exam_column + 1,
                                   value=exam.points_needed_for_6)
                    worksheet.cell(row=amount_header_columns - 3, column=exam_column, value=exam.computation_strategy)
                    worksheet.cell(row=amount_header_columns - 2, column=exam_column, value=exam.name)
                    worksheet.cell(row=amount_header_columns - 2, column=exam_column + 1, value=exam.category)
                    worksheet.cell(row=amount_header_columns - 1, column=exam_column, value="Punkte")
                    worksheet.cell(row=amount_header_columns - 1, column=exam_column + 1, value="Note")

                    # store grading data
                    for row, student in enumerate(self.students, start=amount_header_columns):
                        try:
                            points = exam.points[student]
                        except:
                            points = ""
                        worksheet.cell(row=row, column=exam_column, value=points)
                        try:
                            grade = exam.grades[student]
                        except:
                            grade = ""
                        worksheet.cell(row=row, column=exam_column + 1, value=grade)

            for row in range(2, amount_header_columns - 2):
                worksheet.row_dimensions[row].hidden = True

            logging.info(f"Saved Exam grades to Excel sheet on location{output_name}")
            workbook.save(output_name)


    def create_grade_report(self, output_location: str = None, output_name: str = None,  output_type = "xlsx"):
        """
        creates new report, i.e. computes average grades for each student
        :argument output_location: location where final report should be stored
        :return:
        """

        # TODO merge with export_class_overview to a single export function
        # TODO use this to check my new implementation for errors in unit testing

        # basic test has worked, need to test ugly cases (students did not take exam, ...)

        raise NotImplementedError

        if output_type != "xlsx":
            logging.info(f"Don't know filetype {output_type}, using .xlsx instead")
            output_type = "xlsx"

        if not self.categories:
            raise RuntimeError("Don't have any exams stored, can not generate grade report")

        sum_weights = 0
        for cat in self.categories:
            sum_weights = sum_weights + cat.weight
        if sum_weights != 1:
            raise RuntimeError("Sum of weights must be 1. Aborting")

        logging.info("Creating exam report, calculating final grades")

        total_grades: dict[Student, float] = {}
        for student in self.students:
            exams_per_cat: dict[Category, list[Exam]] = {}
            empty_cat = []
            # collect all exams
            for cat in self.categories:
                exams_taken = []
                if len(cat.exams) == 0:
                    empty_cat.append(cat)
                else:
                    for exam in cat.exams:
                        if student in exam.grades:
                            if exam.grades[student] != -1:
                                exams_taken.append(exam)
                if len(exams_taken) == 0:
                    empty_cat.append(cat)
                else:
                    exams_per_cat[cat] = exams_taken

            if empty_cat:
                logging.info(f"Student {student} did not take all exams necessary for easy computation of grade. Skipping this student, add grade manually")
                continue

            # for all those categories where there is no exam seen, I need to add a scaling factor -- see above

            weighted_per_cat: dict[BaseCategory, float] = {}
            for cat in exams_per_cat:
                # compute "average" grade for each category, according to grading type fixed in category
                weighted_per_cat[cat] = cat.aggregate_grades(student, exams_per_cat[cat])

            # compute rescaling factor, if a student missed all exams from a given category
            # rescale everything: w1, w2, w3, w4 s.t. sum(wi) = 1 ==> if c1 is not present for some student,
            # use w2 + (w1/(4-1), w3 + w1/(4-1), w4 + w1/(4-1)
            # replace w1 by sum(empty_cat.weights) and 4-1 by len(exams_per_cat) - 1
            scaling_factor = 0
            for cat in empty_cat:
                scaling_factor = scaling_factor + cat.weight
            scaling_factor = scaling_factor / (len(exams_per_cat)-1)

            # make sure that I did not screw up
            debug_sum = 0
            for cat in weighted_per_cat:
                debug_sum = debug_sum + cat.weight + scaling_factor
            assert debug_sum == 1, f"Rescaled sum = {debug_sum}\n Skipped cateogries: {[cat.name for cat in empty_cat]}\nTaken categories{[cat.name for cat in exams_taken]}"

            # now, compute final grade for that student
            grade_for_student = 0
            for cat in weighted_per_cat:
                grade_for_student = grade_for_student + (cat.weight + scaling_factor) * weighted_per_cat[cat]

            # TODO ugly assert statement -- how to deal with different max / min grades?
            # TODO ignoring for now, in Switzerland we have 1 - 6. No discussion there

            assert grade_for_student >= 1 and grade_for_student <= 6

            total_grades[student] = grade_for_student

            self.students = sorted(self.students, key=lambda student: student.lastname)

        rounded_grades = {}
        for student in total_grades:
            rounded_grades[student] = round(total_grades[student] * 4) / 4

        if output_location == None:
            logging.info(f"Received no output location for the grade report. Using the default location {self.filename_class_base}")
            output_location = self.filename_class_base
        if not output_location.endswith("/"):
            output_location = output_location + "/"

        if output_name == None:
            output_name = f"{self.name}_report{self.report_id:02}"
            self.report_id = self.report_id + 1
            logging.info(f"Using default output name {output_name}")

        if output_type == "xlsx":
            if not output_name.endswith(".xlsx"):
                if output_name.endswith(".csv"):
                    logging.warning("Got .csv as file ending. Ignoring the user input and choosing .xlsx instead (try changing mode to csv, if you want to store the file as a csv instead)")
                    output_name = output_name[:-4]
                output_name = output_name + ".xlsx"

                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                worksheet.cell(row=1, column=1, value="Nachname")
                worksheet.cell(row=1, column=2, value="Vorname")
                worksheet.cell(row=1, column=3, value="Note Exakt")
                worksheet.cell(row=1, column=4, value="Note Gerundet")

                # write the data
                for row, student in enumerate(self.students, start=2):
                    worksheet.cell(row=row, column=1, value=student.lastname)
                    worksheet.cell(row=row, column=2, value=student.firstname)
                    if student in total_grades:
                        worksheet.cell(row=row, column=3, value=total_grades[student])
                        worksheet.cell(row=row, column=4, value=rounded_grades[student])
                    else:
                        logging.info(f"Skipping student {student}, don't have a grade for them")
                        # row = row - 1

                workbook.save(output_location + output_name)
                logging.info(f"Saved Excel sheet with results to {output_location + output_name}")

        else:
            if output_type == "csv":
                if not output_name.endswith(".csv"):
                    if output_name.endswith(".xlsx"):
                        logging.warning(
                            "Got .xlsx as file ending. Ignoring the user input and choosing .csv instead (try changing mode to xlsx, if you want to store the file as a Excel sheet instead)")
                        output_name = output_name[:-5]
                    output_name = output_name + ".csv"

                    with open(output_location + output_name, "w", newline="") as file:
                        writer = csv.writer(file)

                        # write the headers
                        writer.writerow(["Nachname", "Note exakt", "Note gerundet"])

                        # write the data
                        for i, student in enumerate(self.students):
                            writer.writerow([student.lastname, student.firstname, total_grades[student], rounded_grades[i]])
            else:
                raise RuntimeError(f"Unknown output type {output_type}. Use xlsx or csv, instead")
