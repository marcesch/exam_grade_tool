# This is a sample Python script.
import csv
import logging
import os

import openpyxl
from appdirs import *
# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


from backend.exam import *
from backend.classes import Class
from backend.category import *
import jsonpickle



"""
TODO keep list of classes to index everything -- in GUI, user can then select which class to edit
TODO add functionality to delete old classes
TODO upon startup, load everything from csv
=> sample run:
1. Load list of classes from disk (stored as csv)
2. Let user choose class to be viewed / edited
3. Load class-list, list of exams and exams themselves from stored csv
=> either just load all exams found in a folder or keep index (stored somewhere as csv), probably the latter as I also 
need to store information about the category weights / ...

=> Probably need to create new file for loading / storing stuff. Keep it minimal, loading takes quite a long time...
TODO don't always store everything on disk, only on major changes and on shutdown!

TODO add fucntionality to create exam summary as a PDF or so (average, mean, quartils, standard deviation, ..)
"""

FOLDERPATH = "/home/marcesch/noten/tmp/klassen/"

class Overview:

    def __init__(self):
        self.classes = []
        self.terms = []

        # appdirs takes care of cross-plattform user data storage in according location
        self.application_name = "grade_calculator"
        self.author = "ch.marcesch"
        self.application_folder = user_data_dir(self.application_name, self.author, roaming=False)
        self.config_folder = user_config_dir(self.application_name)

    def save_to_json(self, filename="overview_classes.json"):
        # Combine the provided filename with the application data directory
        file_path = os.path.join(self.application_folder, filename)

        # Serialize the Overview object to a JSON string using jsonpickle
        json_data = jsonpickle.encode(self)

        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # Write the JSON data to the specified file
        with open(file_path, 'w') as json_file:
            json_file.write(json_data)

    @staticmethod
    def load_from_json(filename="overview_classes.json"):
        print(f"=====> Filename {filename}")

        # Combine the provided filename with the application data directory
        application_folder = user_data_dir("grade_calculator", "ch.marcesch", roaming=False)
        file_path = os.path.join(application_folder, filename)

        print(f"[OVERVIEW] Loading data from {file_path}")

        # Read the JSON data from the file
        with open(file_path, 'r') as json_file:
            json_data = json_file.read()

        print(f"[OVERVIEW] Found the following data: \n{json_data}")

        # Deserialize the JSON data using jsonpickle and create an Overview object
        overview = jsonpickle.decode(json_data)
        print(f"[OVERVIEW] After deserialization: \n {overview} \n{overview.classes} ")

        print(f"Created class {overview.classes[0]} with students\n {overview.classes[0].students}")
        print(f"Points for exam {overview.classes[0].exams[0]}: {overview.classes[0].exams[0].points}")
        print(f"Grades for exam {overview.classes[0].exams[0]}: {overview.classes[0].exams[0].grades}")

        return overview


    def save_config(self):
        raise NotImplementedError

    def get_class(self, name, year, term):
        """
        Returns the class object matching the name / year / term ID
        :param name:
        :param year:
        :param term:
        :return: class object
        """

        for class_obj in self.classes:
            if class_obj.name.lower() == name.lower() and class_obj.year == int(year) and class_obj.term.lower() == term.lower():
                return class_obj
        raise RuntimeError(f"Did not find any class matching {name} {term} {year}")

    def get_all_classes_of_period(self, term, year):
        """
        :param term:
        :param year:
        :return: a list of classes matching the term / year
        """

        res = []
        for class_obj in self.classes:
            if (class_obj.term == term) and (class_obj.year == int(year)):
                res.append(class_obj)

        return res

    def return_newest_classes(self, num_semester):
        """
        Returns a list of the last num_semester many seen tuples <term>-<year>. E.g. returns [HS2020, FS2021] if num_semester is 2
        :param num_semester: Lenght of list to be returned
        :return: tuple (term,year) of newest classes in self.classes
        """

        def term_key(cls):
            return (0 if cls == 'FS' else 1, cls)

        term_years = []
        for class_obj in self.classes:
            if (class_obj.term, class_obj.year) in term_years:
                continue
            else:
                term_years.append((class_obj.term, class_obj.year))

        term_years.sort(key=lambda cls: (cls[1], term_key(cls[0])))
        if len(term_years) <= num_semester:
            logging.info(f"Did not find {num_semester} exams, only returning {len(term_years)} many")
            return term_years
        else:
            return term_years[-num_semester:]

    def delete_class(self, class_obj: Class):
        """
        Delete the class object, including all data!
        :param class_obj:
        :return: False if failure to delete
        """

        if class_obj in self.classes:
            self.classes.remove(class_obj)
        else:
            return False


    def add_class(self, name, term, year):
        """
        Add a class object to the overview
        :param name: class name (e.g. 6a)
        :param term:
        :param year:
        :return:
        """

        for class_obj in self.classes:
            if class_obj.name.lower() == name.lower() and class_obj.term.lower() == term.lower() and class_obj.year == year:
                raise RuntimeError(f"[OVERVIEW] Cannot create duplicate class {name} {term} {year}. Consider choosing a new name")

        if not (term.lower() == "hs" or term.lower() == "fs"):
            raise RuntimeError(f"[OVERVIEW] Choose either HS or FS for term")

        class_obj = Class(name, term, year)
        self.classes.append(class_obj)

    def add_students_to_class(self, students: list[tuple[str, str]], class_obj: Class):
        """
        Add interface to add students to class
        :param class_obj:
        :return: list of students that could NOT be added
        """

        list_of_failed_students = class_obj.add_multiple_students(students)
        if list_of_failed_students:
            logging.error(f"[OVERVIEW] Error while creating some students. Most likely they were added a second time: {list_of_failed_students}")
            return list_of_failed_students


    def load_categories_and_exams(self, class_obj: Class, path_folder_exams: str = None):
        """
        :return:
        """

        # TODO replace with an import function

        # TODO I think deprecated, discarding for now
        raise NotImplementedError

        if path_folder_exams == None:
            dirpath = class_obj.filename_base_exam
        else:
            dirpath = path_folder_exams

        logging.info(f"Looking at {dirpath} for class data")

        # TODO include checks on validity / existence of folder

        path = os.path.join(dirpath, "pruefungen.xlsx")

        # Open workbook
        wb = openpyxl.load_workbook(path)
        sheet = wb.active

        # Find row containing "Nachname"
        # TODO I think this is wrong with max_row, using 20 instead.
        # for row in sheet.iter_rows(min_row=1, max_row=1):
        row_index = -1
        for row in sheet.iter_rows(min_row=1, max_row=20):
            for cell in row:
                if cell.value == "Nachname":
                    row_index = cell.row

        if row_index == -1:
            raise RuntimeWarning(
                f"Could not locate the string Nachname in the first row. Please make sure that you did not alter the contents of that file")

        # TODO insert checks if things go wrong
        # Iterate over category rows
        for row in sheet.iter_rows(min_row=2, max_row=row_index - 5, min_col=2, max_col=4):
            category_name = row[0].value
            weight = float(row[1].value)
            grading_type = row[2].value
            category = Category(category_name, class_obj.term, class_obj.name, weight, grading_type)
            class_obj.categories.append(category)

        # Now read in the contents for the categories => iterate over all columns that contain data for exams
        col_index = 3

        while True:
            # read contents for current exam
            cell_min_grade = sheet.cell(row=row_index - 4, column=col_index)
            cell_max_grade = sheet.cell(row=row_index - 4, column=col_index + 1)
            cell_max_points = sheet.cell(row=row_index - 3, column=col_index)
            cell_pts_for_max = sheet.cell(row=row_index - 3, column=col_index + 1)
            cell_cmp_mode = sheet.cell(row=row_index - 2, column=col_index)
            cell_exam_name = sheet.cell(row=row_index - 1, column=col_index)
            cell_exam_cat = sheet.cell(row=row_index - 1, column=col_index + 1)

            # check if there still is data for exam in this column
            if cell_exam_name.value is None or cell_exam_name.value == "":
                break

            # create new exam and add it to correct category
            min_grad = float(cell_min_grade.value)
            max_grade = float(cell_max_grade.value)
            max_points = int(cell_max_points.value)
            pts_for_max = int(cell_pts_for_max.value)
            computation_mode = cell_cmp_mode.value
            exam_name = cell_exam_name.value
            exam_cat = cell_exam_cat.value

            # read in  student data:
            grades_exam = {}
            points_exam = {}
            number_iteration = 0
            for row in sheet.iter_rows(min_row=row_index +1):
                firstname = row[1].value
                lastname = row[0].value
                if firstname == None and lastname == None:
                    break
                student = class_obj.get_student(firstname, lastname)
                if row[col_index] != None:
                    print(f"Here with curr row: {(row[col_index - 1]).value}\npath {path}")
                    points_exam[student] = int((row[col_index-1]).value)
                try:
                    if row[col_index ] != None:
                        grades_exam[student] = float((row[col_index ]).value)
                except Exception as e:
                    print(f"Error for col {col_index+1}\n{e}\n{row[col_index]}")
                # add break condition: stop after 100 iterations
                if number_iteration >= 100:
                    break
                number_iteration += 1



            # TODO remove print
            print(f'Staring to add new exam with grades / points: \n  Grades: {grades_exam}\n Points: {points_exam}')
            # logging.info(f"")

            # create exam object
            class_obj.add_exam(exam_name,
                               exam_cat,
                               max_points,
                               pts_for_max,
                               min_grade=min_grad,
                               max_grade=max_grade,
                               achieved_points=points_exam,
                               achieved_grades=grades_exam,
                               computation_mode=computation_mode)

            col_index += 2


    def store_config(self):
        raise NotImplementedError

    def load_config(self):
        raise NotImplementedError

    def create_import_template(self):
        """
        Create template that can be filled in by user to initialize class
        :return:
        """
        raise NotImplementedError

    def update_semester(self, class_obj: Class):
        """
        Updates a class to the new semester by initializing a new class object with the same categories, students, new name, ...
        :param class_obj:
        :return:
        """
        raise NotImplementedError
