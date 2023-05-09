import csv
import typing
import os
import logging
import shutil
import openpyxl
import time

from category import Category
from exam import Exam
from student import Student

FOLDERPATH = "./tmp/klassen/"

""""
TODO: sanitize / ... students names... what if there are sume ugly shit thigns like accents etc?
=> might have to slightly alter the contains_student function etc. (always compare sanitized version)

=> add possibility to import old examples (e.g. for categories etc., otherwise have to do that everytime again)


TODO: when should I store stuff?

TODO cna python also handle xlsx files? might be a little bit more convenient than using csv... e.g. when creating report

"""


class Class:
    def __init__(self, name: str, term: str, year: int):
        self.students: list[Student] = []
        self.categories: list[Category] = []
        self.name: str = name
        if (term.lower() == "hs"):
            self.term: str = term.upper()
        else:
            if (term.lower() == "fs"):
                self.term: str = term.upper()
            else:
                raise RuntimeError("Please choose either HS or FS")
        self.year: int = year
        self.filename_class_base = ""
        self.filename_class = ""
        self.filename_base_examfilename_base_exam = ""
        self.update_filenames()
        self.folder_shadow = FOLDERPATH + ".trash/"
        self.report_id = 0
        self.supported_filetypes = ["xlsx", "csv"]

        if not os.path.exists(self.filename_base_exam):
            os.makedirs(self.filename_base_exam)
        else:
            logging.warning(f"Directory '{self.filename_base_exam}' already exists.")

    def __str__(self):
        return f"{self.name} {self.term.upper()}{self.year}"

    def __repr__(self):
        return f"C-{self.name} {self.term.upper()}{self.year}"


    def update_name(self, new_name: str):
        self.name = new_name
        self.update_filenames()
        # TODO rename stored file containing list of class, otherwise name gets "rewritten" once program gets loaded.

    def update_filenames(self):
        """
        after seen a change (e.g. year / term / ...), need to adjust the filename
        :return:
        """
        # TODO ask selina for her preferences regarding layout of files
        # TODO remove old files
        classname_complete = str(self.year) + "_" + self.term.upper() + "_" + self.name
        self.filename_class_base = FOLDERPATH + classname_complete
        self.filename_class = self.filename_class_base + ".csv"
        # or use self.filename_class = self.filename_class_base + "/" + classname_complete + ".csv"
        self.filename_base_exam = self.filename_class_base + "/pruefungen/"

    def update_semester(self):
        """
        :return:

        updates the semester by one term
        creates new file for grades
        """
        self.report_id = 0
        oldterm = self.term
        oldyear = self.year
        if self.term == "hs":
            self.year = self.year + 1
            self.term = "fs"
        else:
            self.term = "hs"
            logging.info("Consider updating the name of the class! It has likely changed, e.g. from 3a to 4a")

        logging.info(f"Updating term for class {self.name} from {oldterm} {oldyear} to {self.term} {self.year}")
        logging.info(f"Removing old class list, updating with new classlist")
        old_filename = self.filename_class
        # TODO also make new files for class name, deleting the old ones
        self.update_filenames()
        os.remove(old_filename)
        self.store_to_database()

        # make new base folder for exams
        if not os.path.exists(self.filename_base_exam):
            os.makedirs(self.filename_base_exam)
        else:
            logging.error(f"Directory '{self.filename_base_exam}' already exists.")

        # Reuse cateogries from previous semester, too.
        old_categories = self.categories
        self.categories = []
        self.initialize_categories_from_old(old_categories)



    def store_to_database(self):
        """
        Stores the old class file to the trash-folder and stores an updated version
        :return:
        """
        students_old = []
        # load "old" student's database for comparison
        # Read the list of students from the CSV file
        try:
            with open(self.filename_class, "r") as csvfile:
                reader = csv.reader(csvfile)
                # Skip the header row
                next(reader)
                # Iterate over the rows in the CSV file and create a new Student object for each row
                for row in reader:
                    student = Student(row[0], row[1])
                    students_old.append(student)

            # find difference between the two files
            students_removed = []
            for student in students_old:
                b, student_diff = self.contains_student(student.firstname, student.lastname)
                if b:
                    students_removed.append(student_diff)

            if len(students_removed) == 0 and len(students_old) == len(self.students):
                print("The class has not changed. Won't store anything")
                return

            if len(students_removed) != 0:
                print(
                    f"REMOVING the following students from the database. If this was an error, you can find the old file in {self.folder_shadow}")
                for student in students_removed:
                    print(f"Removing student {student.firstname} {student.lastname}")

            # move "old" student's database to trash-folder
            logging.info("Moving old database to trash folder")
            shutil.move(self.filename_class, f"{self.folder_shadow}{time.time()/60}_{self.filename_class}")
        except:
            logging.info(f"No old class list found on disk")

        # store "new" / current database under folderpath
        # Write the list of students to a CSV file
        with open(self.filename_class, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            # Write the header row to the CSV file
            writer.writerow(["Nachname", "Vorname"])
            # Write each student's first name and last name to the CSV file
            for student in self.students:
                writer.writerow([student.lastname, student.firstname])

    def get_student(self, first: str, last: str):
        """
        Really just a wrapper function for contains_student
        :param first:
        :param last:
        :return:  Student object corresponding to the firstname / lastname combination
        """
        _, student = self.contains_student(first, last, return_student=True)
        return student

    def contains_student(self, first: str, last: str, return_student=False):
        """
        Slow and unelegant, but hey
        Allows only searching for one of the names (surname or firstname), this is a little bit more convenient
        :param return_student:  can let the function also return the Student object (used for deletion)
        :param first:
        :param last:
        :return:
        """
        if first == None and last == None:
            raise RuntimeError("Must give at least one name")
        res = False
        res_student = None
        for students_contained in self.students:
            if students_contained.firstname == first and students_contained.lastname == last:
                res = True
                res_student = students_contained
            if first == None and students_contained.lastname == last:
                res = True
                res_student = students_contained
            if first == students_contained.firstname and last == None:
                res = True
                res_student = students_contained

        if return_student:
            return res, res_student
        return res

    def initialize_new_class(self, students: list[dict[str, str]]):
        """
        creates new class from scratch and stores the data of the students as a csv.
        :param
        :return:
        """

        for student in students:
            if self.contains_student(student["firstname"], student["lastname"]):
                print(f"Warning! Received the same student a second time! Won't add anything")
                # TODO find some reasonable way to ping that back to GUI (do you really want to do that? or so)
            else:
                self.students.append(Student(student["firstname"], student["lastname"]))

        # store the class as a csv
        # => that function checks whether old file already exists
        self.store_to_database()

    def delete_class(self):
        """
        deltes class AND ALL ITS FILES
        :return:
        """
        # TODO could also just move it to the trash folder -- need to decide that later
        logging.warning(f"Removing all the files for class {self.name} {self.term} {self.year}. This action cannot be reverted.")
        os.remove(self.filename_class_base)

        # put the old folder to the trash folder
        shutil.move(self.filename_base_exam, f"{self.folder_shadow}{time.time() / 60}_{self.filename_class}")

        # remove the folder containing this classes exam
        for item in os.listdir(self.filename_base_exam):
            item_path = os.path.join(self.filename_base_exam, item)
            if os.path.isfile(item_path):
                # if the item is a file, delete it
                os.remove(item_path)
            elif os.path.isdir(item_path):
                # if the item is a directory, recursively delete it and all its contents
                shutil.rmtree(item_path)

    def empty_trash(self):
        """
        removes everything from the trash folder, raising tons of warnings

        :return:
        """
        logging.warning("DELETING the entire contents of the trash folder")

        for item in os.listdir(self.folder_shadow):
            item_path = os.path.join(self.folder_shadow, item)
            if os.path.isfile(item_path):
                # if the item is a file, delete it
                os.remove(item_path)
            elif os.path.isdir(item_path):
                # if the item is a directory, recursively delete it and all its contents
                shutil.rmtree(item_path)

    def add_student(self, first: str, last: str):
        """
        :return:
        adds new student
        """
        if self.contains_student(first, last):
            logging.warning(f"Class already contains student {first.capitalize()} {last.capitalize()}, skipping")
        else:
            self.students.append(Student(first, last))
        self.store_to_database()

    def remove_student(self, first: str, last: str):
        """
        what to do with the grades???
        :param student:
        :return:
        """
        b, student_in_db = self.contains_student(first, last, return_student=True)
        if not b or student_in_db == None:
            raise RuntimeWarning(f"Could not find student {first} {last} in my database. Did you mistype?")

        self.students.remove(student_in_db)
        self.store_to_database()

    def contains_category(self, name: str):
        for cat in self.categories:
            if cat.name == name:
                return True
        return False

    def find_category(self, name: str):
        for cat in self.categories:
            if cat.name == name:
                return cat

        raise RuntimeError(f"Could not find that Category in the list {self.categories}")

    def remove_category(self, cat: str):
        if self.contains_category(cat):
            category = self.find_category(cat)
            self.categories.remove(category)
        else:
            raise RuntimeError(f"Could not find that category in the list {self.categories}")

    def add_category(self, cat: dict[str: str]):
        logging.info("Adding category")
        # TODO check if categroy already exists
        grading_type = cat["grading_type"] if "grading_type" in cat else "default"
        self.categories.append(Category(name=cat["name"], term=f"{self.year}{self.term.upper()}", classname=self.name,
                                        weight=cat["weight"], grading_type=grading_type))

    def initialize_categories_from_old(self, exam_categories: list[Category]):
        for cat in exam_categories:
            self.categories.append(Category(cat.name, f"{self.year}{self.term.upper()}", self.name, cat.weight, cat.grading_type))

    def initialize_categories(self, exam_categories: list[dict[str: typing.Any]]):
        """
        Not quite sure how exactly to do that part -- It might be reasonable to use a wrapper function that checks that
        the weight of the categories always sum to 1 (invariant regarding the correctness) instead of letting user create
        categories individually

        Downside: need to create all of them at once, always. But I think, GUI can help me here to make it reasonably nice to use

        :return:
        """

        # TODO deal with same keys -- different entries -> raise exception or so

        # TODO check for missing weights (or let caller handle that)

        logging.info(f"Initializing list of categories")
        for cat in exam_categories:
            grading_type = cat["grading_type"] if "grading_type" in cat else "default"
            self.categories.append(Category(name=cat["name"], term=f"{self.year}{self.term.upper()}", classname=self.name, weight=cat["weight"], grading_type=grading_type))

    def upadte_categories(self, exam_categories: list[dict[str: str]]):
        """
        Not quite sure how exactly to do that part -- It might be reasonable to use a wrapper function that checks that
        the weight of the categories always sum to 1 (invariant regarding the correctness) instead of letting user create
        categories individually

        Downside: need to create all of them at once, always. But I think, GUI can help me here to make it reasonably nice to use

        :argument exam_categories: list of exam-category info. Use keys "name", "weight" and "grading_type"
        :return:
        """

        # Check that sum always stays one -- what argumend should I get?
        # maybe list with updated categories -- check that every existent category still exists (otherwise, exams get deleted!!!)

        new_names = []
        for cat in exam_categories:
            new_names.append(cat["name"])

        for old_cat in self.categories:
            if old_cat.name not in new_names:
                if len(old_cat.exams) != 0:
                    raise RuntimeError(f"Cannot delete category {old_cat} as there is an exam in that Category.")

        logging.info(f"Updating list of categories")
        for cat in exam_categories:
            if self.contains_category(cat["name"]):
                cat_in_list = self.find_category(cat["name"])
                cat_in_list.weight = cat["weight"]
                if "grading_type" in cat:
                    cat_in_list.grading_type = cat["grading_type"]
            else:
                self.add_category(cat)


    def add_exam(self, exam_name: str,
                 exam_category: str,
                 max_points: int,
                 points_needed_for_6: int = None,
                 min_grade=1,
                 max_grade=6,
                 achieved_points = None,
                 achieved_grades = None,
                 computation_mode=None):
        """

        :param exam_name: e.g. Redaction 1
        :param exam_category: e.g. redaction, orale, ...
        :param max_points: maximum possible points
        :param points_needed_for_6:
        :param min_grade:
        :param max_grade:
        :return:

        TODO make sure to support whacky characters in name
        TODO store .csv file for that exam notes
        TODO store csv with full names
        TODO use special number like -1 to signify "don't count this exam" => what if category does not contain any exam for that studen?
        """

        # TODO sanitize user inputs
        # TODO make sure that exam names are unique

        # check if category is present:
        category = None
        for cat in self.categories:
            if cat.name == exam_category:
                category = cat
                break
        if category is None:
            logging.warning(f"Did not find category. Creating new one. NEED TO SET WEIGHT MANUALLY!")
            category = Category(exam_category, self.term, self.name, weight=0)
            self.categories.append(category)

        new_exam = category.add_exam(exam_name,
                                     term=f"{self.term.upper()}_{self.year}",
                                     classname=self.name,
                                     max_points=max_points,
                                     points_needed_for_6=points_needed_for_6,
                                     min_grade=min_grade,
                                     max_grade=max_grade,
                                     achieved_points=achieved_points,
                                     achieved_grades=achieved_grades,
                                     computation_mode=computation_mode)

        # computing grades
        new_exam.compute_grades()
        # for cat in self.categories:
        #     for exam in cat.exams:
        #         # print(f"Calling compute_grades on {exam}")
        #         if exam.grades != None:
        #             exam.compute_grades()
        #         else:
        #             logging.warning(f"Cannot compute grades for {exam}; exam.grades is None")

        # making sure to store that exam in Excel file
        self.store_exams()

    def store_exams(self, filetype="xlsx"):
        """
        Stores all exams to disk. Most important function, as this allows manual overwrite
        See the layout_gradefiles for details on what goes where, ...
        :param filetype:
        :return:
        """

        # TODO sort students according to last name
        # sort students according to lastname
        # TODO can also support firstname, theoretically
        self.students = sorted(self.students, key=lambda student: student.lastname)

        # sanitize file type
        # if filetype not in self.supported_filetypes:
        if filetype != "xlsx":
            logging.warning(f"Don't know that filetype {filetype}. Using xlsx instead")
            filetype = "xlsx"

        output_name = os.path.join(self.filename_base_exam, "pruefungen.xlsx")
        # check if an old exam with that name already exists - if yes, move it to trash
        # logging.info(f"Storing {exam.name}_{exam.term} to {output_name}.{filetype}")
        # print(f"Storing {exam.name}_{exam.term} to {output_name}.{filetype}")

        # if os.path.exists(output_name):
        #     logging.warning("Found old file containing the exams. Moving it to the .trash folder")
        #     shutil.move(output_name, f"{self.folder_shadow}{time.time()/60}_{output_name}")

        # store the exam to disk, given the chosen mode
        # TODO also store max points, points for 6 in exam
        if filetype == "xlsx":
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # reserve some space on top of Excel file (for class name, weights of categories, ...)
            # For loading, search for "Nachname" or so to get this index.

            # set number of headers, accoring to layout in .ods file
            amount_header_columns = 1 + 1 + len(self.categories) + 5

            worksheet.cell(row=1, column=1, value="Klassenname")
            worksheet.cell(row=1, column=2, value=self.name)

            worksheet.cell(row=2, column=1, value="Categories")
            worksheet.cell(row=2, column=2, value="Category name")
            worksheet.cell(row=2, column=3, value="Category weight")
            worksheet.cell(row=2, column=4, value="Grading type")
            for i, cat in enumerate(self.categories):
                worksheet.cell(row=2 + i, column=2, value=cat.name)
                worksheet.cell(row=2 + i, column=3, value=cat.weight)
                worksheet.cell(row=2 + i, column=4, value=cat.grading_type)
                worksheet.row_dimensions[2+i].hidden = True

            worksheet.cell(row=amount_header_columns-1, column=1, value="Nachname")
            worksheet.cell(row=amount_header_columns-1, column=2, value="Vorname")
            for row, student in enumerate(self.students, start=amount_header_columns):
                worksheet.cell(row=row, column=1, value=student.lastname)
                worksheet.cell(row=row, column=2, value=student.firstname)

            column_iterator = 0
            for cat_column, cat in enumerate(self.categories):
                for exam_column, exam in enumerate(cat.exams):
                    # store exam headers
                    exam_column = 2*column_iterator + 3
                    column_iterator = column_iterator + 1
                    worksheet.cell(row=amount_header_columns-5, column=exam_column, value=exam.min_grade)
                    worksheet.cell(row=amount_header_columns-5, column=exam_column+1, value=exam.max_grade)
                    worksheet.cell(row=amount_header_columns-4, column=exam_column, value=exam.max_points)
                    worksheet.cell(row=amount_header_columns-4, column=exam_column+1, value=exam.points_needed_for_6)
                    worksheet.cell(row=amount_header_columns-3, column=exam_column, value=exam.computation_strategy)
                    worksheet.cell(row=amount_header_columns-2, column=exam_column, value=exam.name)
                    worksheet.cell(row=amount_header_columns-2, column=exam_column+1, value=exam.category)
                    worksheet.cell(row=amount_header_columns-1, column=exam_column, value="Punkte")
                    worksheet.cell(row=amount_header_columns-1, column=exam_column+1, value="Note")

                    # store grading data
                    for row, student in enumerate(self.students, start=amount_header_columns):
                        try:
                            points = exam.points[student]
                        except:
                            points = ""
                        worksheet.cell(row=row, column=exam_column, value=points)
                        try:
                            grade = exam.grades[student]
                        except:
                            grade = ""
                        worksheet.cell(row=row, column=exam_column+1, value=grade)

            for row in range(2, amount_header_columns-2):
                worksheet.row_dimensions[row].hidden = True

            logging.info(f"Saved Exam grades to Excel sheet on location{output_name}")
            workbook.save(output_name)

    def update_grade(self, exam: Exam, students: list[Student], new_grades: dict[Student, float], new_points: dict[Student, float], computation_mode = "linear"):
        """
        Lets a user update a student's grade, either by letting the program recompute the grade based on their points or manually
        overwriting everything by setting the grades. Gives priority to recalculating stuff by my program
        :param exam:
        :param student:
        :return:
        """

        for student in students:
            if student not in self.students:
                raise RuntimeError(f"Don't know that student {student}. Did you forget to add it to the class {self.name}?")
            # give precedence for letting my program recalculating instead of setting manual grades
            if student in new_points:
                exam.grades[student] = exam.compute_single_grade(new_points[student], computation_mode)
            else:
                if student in new_grades:
                    exam.grades[student] = new_grades[student]
                else:
                    raise logging.error(f"Could not set grade for student {student.firstname} {student.lastname} as they are not in either provided list")

        # store exam to disk
        self.store_exams(exam)



    def create_grade_report(self, output_location: str = None, output_name: str = None,  output_type = "xlsx"):
        """
        creates new report, i.e. computes average grades for each student
        :argument output_location: location where final report should be stored
        :return:
        """

        # TODO thoroughly test this function!!!
        # basic test has worked, need to test ugly cases (students did not take exam, ...)


        if output_type != "xlsx":
            logging.info(f"Don't know filetype {output_type}, using .xlsx instead")
            output_type = "xlsx"

        if not self.categories:
            raise RuntimeError("Don't have any exams stored, can not generate grade report")

        sum_weights = 0
        for cat in self.categories:
            sum_weights = sum_weights + cat.weight
        if sum_weights != 1:
            raise RuntimeError("Sum of weights must be 1. Aborting")

        logging.info("Creating exam report, calculating final grades")

        total_grades: dict[Student, float] = {}
        for student in self.students:
            exams_per_cat: dict[Category, list[Exam]] = {}
            empty_cat = []
            # collect all exams
            for cat in self.categories:
                exams_taken = []
                if len(cat.exams) == 0:
                    empty_cat.append(cat)
                else:
                    for exam in cat.exams:
                        if student in exam.grades:
                            if exam.grades[student] != -1:
                                exams_taken.append(exam)
                if len(exams_taken) == 0:
                    empty_cat.append(cat)
                else:
                    exams_per_cat[cat] = exams_taken

            if empty_cat:
                logging.info(f"Student {student} did not take all exams necessary for easy computation of grade. Skipping this student, add grade manually")
                continue

            # for all those categories where there is no exam seen, I need to add a scaling factor -- see above

            weighted_per_cat: dict[Category, float] = {}
            for cat in exams_per_cat:
                # compute "average" grade for each category, according to grading type fixed in category
                weighted_per_cat[cat] = cat.aggregate_grades(student, exams_per_cat[cat])

            # compute rescaling factor, if a student missed all exams from a given category
            # rescale everything: w1, w2, w3, w4 s.t. sum(wi) = 1 ==> if c1 is not present for some student,
            # use w2 + (w1/(4-1), w3 + w1/(4-1), w4 + w1/(4-1)
            # replace w1 by sum(empty_cat.weights) and 4-1 by len(exams_per_cat) - 1
            scaling_factor = 0
            for cat in empty_cat:
                scaling_factor = scaling_factor + cat.weight
            scaling_factor = scaling_factor / (len(exams_per_cat)-1)

            # make sure that I did not screw up
            debug_sum = 0
            for cat in weighted_per_cat:
                debug_sum = debug_sum + cat.weight + scaling_factor
            assert debug_sum == 1, f"Rescaled sum = {debug_sum}\n Skipped cateogries: {[cat.name for cat in empty_cat]}\nTaken categories{[cat.name for cat in exams_taken]}"

            # now, compute final grade for that student
            grade_for_student = 0
            for cat in weighted_per_cat:
                grade_for_student = grade_for_student + (cat.weight + scaling_factor) * weighted_per_cat[cat]

            # TODO ugly assert statement -- how to deal with different max / min grades?
            # TODO ignoring for now, in Switzerland we have 1 - 6. No discussion there

            assert grade_for_student >= 1 and grade_for_student <= 6

            total_grades[student] = grade_for_student

            self.students = sorted(self.students, key=lambda student: student.lastname)

        rounded_grades = {}
        for student in total_grades:
            rounded_grades[student] = round(total_grades[student] * 4) / 4

        if output_location == None:
            logging.info(f"Received no output location for the grade report. Using the default location {self.filename_class_base}")
            output_location = self.filename_class_base
        if not output_location.endswith("/"):
            output_location = output_location + "/"

        if output_name == None:
            output_name = f"{self.name}_report{self.report_id:02}"
            self.report_id = self.report_id + 1
            logging.info(f"Using default output name {output_name}")

        if output_type == "xlsx":
            if not output_name.endswith(".xlsx"):
                if output_name.endswith(".csv"):
                    logging.warning("Got .csv as file ending. Ignoring the user input and choosing .xlsx instead (try changing mode to csv, if you want to store the file as a csv instead)")
                    output_name = output_name[:-4]
                output_name = output_name + ".xlsx"

                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                worksheet.cell(row=1, column=1, value="Nachname")
                worksheet.cell(row=1, column=2, value="Vorname")
                worksheet.cell(row=1, column=3, value="Note Exakt")
                worksheet.cell(row=1, column=4, value="Note Gerundet")

                # write the data
                for row, student in enumerate(self.students, start=2):
                    worksheet.cell(row=row, column=1, value=student.lastname)
                    worksheet.cell(row=row, column=2, value=student.firstname)
                    if student in total_grades:
                        worksheet.cell(row=row, column=3, value=total_grades[student])
                        worksheet.cell(row=row, column=4, value=rounded_grades[student])
                    else:
                        logging.info(f"Skipping student {student}, don't have a grade for them")
                        # row = row - 1

                workbook.save(output_location + output_name)
                logging.info(f"Saved Excel sheet with results to {output_location + output_name}")

        else:
            if output_type == "csv":
                if not output_name.endswith(".csv"):
                    if output_name.endswith(".xlsx"):
                        logging.warning(
                            "Got .xlsx as file ending. Ignoring the user input and choosing .csv instead (try changing mode to xlsx, if you want to store the file as a Excel sheet instead)")
                        output_name = output_name[:-5]
                    output_name = output_name + ".csv"

                    with open(output_location + output_name, "w", newline="") as file:
                        writer = csv.writer(file)

                        # write the headers
                        writer.writerow(["Nachname", "Note exakt", "Note gerundet"])

                        # write the data
                        for i, student in enumerate(self.students):
                            writer.writerow([student.lastname, student.firstname, total_grades[student], rounded_grades[i]])
            else:
                raise RuntimeError(f"Unknown output type {output_type}. Use xlsx or csv, instead")
